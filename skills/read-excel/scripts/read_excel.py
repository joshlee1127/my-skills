#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""快速讀取任意 .xlsx／.xlsm 檔案的內容，不假設任何欄位結構。

用法：
    python read_excel.py <file.xlsx> --list-sheets
    python read_excel.py <file.xlsx> [--sheet <name>] [--format md|json] [--out <path>]
    python read_excel.py <file.xlsx> --sheet <name> --range A1:D50 --header-row 2
    python read_excel.py <file.xlsx> --fast --max-rows 2000   # 大檔案不需要合併儲存格展開時用

Windows 提醒：stdout 一律以 UTF-8 位元組輸出，管線接給另一支程式容易被 cp950 解碼成亂碼，
需要接續處理就用 `--out` 直接寫檔（跟 create-agents 技能的慣例一致）。

設計重點：
- 預設用一般模式載入（`data_only=True`），會展開合併儲存格——只填在左上角的值，
  沿用到被合併蓋住的其他儲存格，避免看起來像資料缺漏。這是手工維護的 Excel 最常見的
  「明明有資料，讀出來卻是空的」問題來源。
- `--fast` 改用 openpyxl 的 read_only 模式，讀取大檔案明顯更快，但代價是拿不到合併儲存格
  資訊（openpyxl 的限制），所以合併儲存格會維持原樣（只有左上角有值）。檔案很大、
  又確定沒有合併儲存格或不在意時才用。
- 不做欄位語意判斷（不猜哪欄是「項目分類」之類）——這支腳本只管「把格線資料忠實搬出來」，
  語意需要另外判讀的估算表交給 create-agents 技能的 read_estimate.py。
- 不靜默截斷：超過 --max-rows 只會截斷輸出，但一定會在結果裡註明實際列數與截斷位置。
"""

import argparse
import json
import re
import sys

try:
    import openpyxl
    from openpyxl.utils.cell import range_boundaries
except ImportError:  # pragma: no cover
    sys.exit("需要 openpyxl：pip install openpyxl")

DEFAULT_MAX_ROWS = 500


def emit(text, stream=None):
    """以 UTF-8 位元組輸出，避開 Windows 主控台預設 cp950 把中文寫成亂碼。"""
    stream = stream or sys.stderr
    stream.buffer.write((text + "\n").encode("utf-8"))
    stream.buffer.flush()


def die(message):
    emit(message)
    sys.exit(1)


def cell_text(value):
    """把儲存格值轉成可讀字串；整數值的 float（Excel 常見）去掉多餘的 .0。"""
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def escape_cell(text):
    """儲存格內容偶爾含有 | 或換行，會把 Markdown 表格撐破。"""
    return text.replace("|", "\\|").replace("\n", " ")


def load_workbook(path, fast):
    return openpyxl.load_workbook(path, read_only=fast, data_only=True)


def list_sheets(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    result = []
    for ws in wb.worksheets:
        result.append({
            "name": ws.title,
            "max_row": ws.max_row,
            "max_col": ws.max_column,
            "merged_ranges": len(list(ws.merged_cells.ranges)),
        })
    return result


def merge_fill_map(ws):
    """回傳 {(row, col): 該合併範圍左上角的值}，讓被合併蓋住的儲存格也能讀到值。"""
    fill = {}
    for merged_range in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        top_left = ws.cell(row=min_row, column=min_col).value
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                fill[(r, c)] = top_left
    return fill


def read_grid(path, sheet_name, cell_range, fast):
    wb = load_workbook(path, fast)
    ws = wb[sheet_name] if sheet_name else wb.worksheets[0]

    fill = {} if fast else merge_fill_map(ws)

    min_col, min_row, max_col, max_row = (1, 1, ws.max_column, ws.max_row)
    if cell_range:
        try:
            min_col, min_row, max_col, max_row = range_boundaries(cell_range)
        except ValueError:
            die("無法解析 --range %r，格式需為 A1:D50 這種 Excel 範圍表示法。" % cell_range)

    grid = []
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        values = []
        for cell in row:
            value = cell.value
            if value is None and (cell.row, cell.column) in fill:
                value = fill[(cell.row, cell.column)]
            values.append(cell_text(value))
        grid.append(values)

    return ws.title, grid


def split_header(grid, header_row, no_header):
    """回傳 (header_texts_or_None, body_rows)。header_row 是 1-based，相對於 grid 的第一列。"""
    if no_header:
        return None, grid
    idx = (header_row or 1) - 1
    if idx < 0 or idx >= len(grid):
        die("--header-row 超出讀到的列數範圍。")
    header = grid[idx]
    body = grid[idx + 1:]
    return header, body


def dedupe_header(header):
    """標題列有空白或重複欄名時補上欄序，避免 JSON 的 key 互相覆蓋。"""
    seen = {}
    out = []
    for i, name in enumerate(header):
        label = name or f"col{i + 1}"
        if label in seen:
            seen[label] += 1
            label = f"{label}_{seen[label]}"
        else:
            seen[label] = 0
        out.append(label)
    return out


def truncate(rows, max_rows):
    if max_rows is None or len(rows) <= max_rows:
        return rows, 0
    return rows[:max_rows], len(rows) - max_rows


def to_markdown(source, sheet_title, header, body, omitted):
    lines = [f"# {source}（工作表：{sheet_title}）", ""]
    if header:
        lines.append("| " + " | ".join(escape_cell(h) for h in header) + " |")
        lines.append("|" + "---|" * len(header))
    for row in body:
        lines.append("| " + " | ".join(escape_cell(v) for v in row) + " |")
    if omitted:
        lines.append("")
        lines.append(f"> ⚠️ 還有 {omitted} 列未顯示，用 --max-rows 或 --range 調整範圍。")
    return "\n".join(lines)


def to_json_payload(source, sheet_title, header, body, omitted):
    if header:
        keys = dedupe_header(header)
        rows = [dict(zip(keys, row)) for row in body]
    else:
        rows = body
    return {
        "source": source,
        "sheet": sheet_title,
        "header": header,
        "row_count": len(body),
        "omitted_rows": omitted,
        "rows": rows,
    }


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("path")
    ap.add_argument("--list-sheets", action="store_true", help="只列出工作表名稱與大小，不解析內容")
    ap.add_argument("--sheet", help="工作表名稱，預設第一個")
    ap.add_argument("--range", dest="cell_range", help="限制讀取範圍，例如 A1:D50")
    ap.add_argument("--header-row", type=int, help="標題列（1-based，相對於讀到的範圍），預設第 1 列")
    ap.add_argument("--no-header", action="store_true", help="不把任何列當標題，輸出原始格線")
    ap.add_argument("--format", choices=["json", "md"], default="md")
    ap.add_argument("--max-rows", type=int, default=DEFAULT_MAX_ROWS,
                     help="輸出的最大資料列數，0 表示不限制（預設 %d）" % DEFAULT_MAX_ROWS)
    ap.add_argument("--fast", action="store_true",
                     help="用 read_only 模式讀取，大檔案更快，但不會展開合併儲存格")
    ap.add_argument("--out", help="輸出檔路徑（UTF-8）。要接續處理時用這個，避免管線編碼問題。")
    args = ap.parse_args()

    if args.list_sheets:
        sheets = list_sheets(args.path)
        out = json.dumps(sheets, ensure_ascii=False, indent=2)
    else:
        sheet_title, grid = read_grid(args.path, args.sheet, args.cell_range, args.fast)
        header, body = split_header(grid, args.header_row, args.no_header)
        max_rows = None if args.max_rows == 0 else args.max_rows
        body, omitted = truncate(body, max_rows)
        if args.format == "json":
            payload = to_json_payload(args.path, sheet_title, header, body, omitted)
            out = json.dumps(payload, ensure_ascii=False, indent=2)
        else:
            out = to_markdown(args.path, sheet_title, header, body, omitted)

    if args.out:
        with open(args.out, "w", encoding="utf-8") as f:
            f.write(out + "\n")
        emit("wrote %s" % args.out, sys.stdout)
        return

    sys.stdout.buffer.write(out.encode("utf-8"))
    sys.stdout.buffer.write(b"\n")


if __name__ == "__main__":
    main()
