#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""把工時估算表（xlsx）解析成結構化資料，供撰寫 AGENTS.md 使用。

用法：
    python read_estimate.py <estimate.xlsx> [--format json|md] [--sheet <name>] [--out <path>]

預設輸出 JSON（UTF-8）到 stdout；`--format md` 輸出人類可讀的大綱，方便直接貼進思考流程。

Windows 提醒：stdout 一律以 UTF-8 位元組輸出，但用管線接給另一支程式時，
對方可能以 cp950 解碼而讀成亂碼／JSON 解析失敗。要接續處理時請用 `--out` 直接寫檔。

設計重點：
- 工時估算表常見的三個坑：標題列不在第 1 列、「項目分類」只寫在該區塊第一列（合併儲存格）、
  每個區塊尾端有「小計」列。這支腳本都會處理，避免每次用技能時重新寫一份解析程式。
- 欄位名稱以關鍵字模糊比對，所以欄位順序不同、多了「作業數量」之類的欄位也不會壞掉。
- 手工維護的估算表常有資料打錯欄（例如標題「驗收項目」在 K 欄、內容卻打在 L 欄）。
  這種資料不會被丟掉，而是收進 item["extra"] 並在 warnings 中示警，讓使用者判斷該歸到哪一欄。
"""

import argparse
import json
import re
import sys

try:
    import openpyxl
except ImportError:  # pragma: no cover
    sys.exit("需要 openpyxl：pip install openpyxl")


# 欄位關鍵字 → 正規化欄名。比對時取第一個命中的關鍵字。
COLUMN_PATTERNS = [
    ("category", ["項目分類", "分類", "作業分類"]),
    ("no", ["項次", "編號", "序號"]),
    ("task", ["作業項目", "工作項目", "項目名稱"]),
    ("qty", ["作業數量", "數量"]),
    ("unit", ["作業單位", "單位"]),
    ("note", ["項目備註", "備註", "說明"]),
    ("acceptance", ["驗收項目", "驗收標準", "交付項目"]),
]

# 工時欄位另外處理：欄名含 L1/L2/工時/人天/人時 的都收進 hours。
HOURS_PATTERN = re.compile(r"(L\d|工時|人時|人天|時數)")

SUBTOTAL_WORDS = ("小計", "合計", "總計", "小　計", "Subtotal", "Total")


def emit(text, stream=None):
    """以 UTF-8 位元組輸出，避開 Windows 主控台預設 cp950 把中文寫成亂碼。"""
    stream = stream or sys.stderr
    stream.buffer.write((text + "\n").encode("utf-8"))
    stream.buffer.flush()


def die(message):
    emit(message)
    sys.exit(1)


def cell_text(value):
    if value is None:
        return ""
    return str(value).strip()


def normalize_multiline(text):
    """儲存格內的換行在估算表裡代表「條列」，保留成 list 比壓成一行有用。"""
    parts = [p.strip() for p in re.split(r"[\r\n]+", text) if p.strip()]
    return parts


def find_header_row(ws, max_scan=15):
    """找出標題列：第一列同時出現「作業項目」類與其他已知欄位關鍵字的列。

    關鍵字直接取自 COLUMN_PATTERNS/HOURS_PATTERN，避免這裡另外維護一份字面
    清單、跟欄位比對邏輯本身的關鍵字表脫節（例如新增了同義詞卻漏改這裡）。
    """
    task_keywords = next(kws for key, kws in COLUMN_PATTERNS if key == "task")
    other_keywords = [kw for key, kws in COLUMN_PATTERNS if key != "task" for kw in kws]
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, max_scan)):
        texts = [cell_text(c.value) for c in row]
        joined = " ".join(texts)
        if any(k in joined for k in task_keywords) and (
            any(k in joined for k in other_keywords) or HOURS_PATTERN.search(joined)
        ):
            return row[0].row, texts
    return None, None


def map_columns(header_texts):
    """回傳 {正規化欄名: 欄索引(0-based)} 與 hours 欄清單。"""
    mapping = {}
    hours_cols = []
    for idx, text in enumerate(header_texts):
        if not text:
            continue
        flat = text.replace(" ", "")
        matched = False
        for key, keywords in COLUMN_PATTERNS:
            if key in mapping:
                continue
            if any(kw in flat for kw in keywords):
                mapping[key] = idx
                matched = True
                break
        if not matched and HOURS_PATTERN.search(flat):
            hours_cols.append((idx, text))
    return mapping, hours_cols


def parse(path, sheet_name=None):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.worksheets[0]

    header_row, header_texts = find_header_row(ws)
    if header_row is None:
        die(
            "找不到標題列（需含「作業項目」與「項目分類/項次/備註/工時」之一）。"
            "請用 --sheet 指定工作表，或確認這份檔案是工時估算表。"
        )

    cols, hours_cols = map_columns(header_texts)
    if "task" not in cols:
        die("標題列缺少「作業項目」欄，無法解析。")

    def get(row, key):
        idx = cols.get(key)
        if idx is None or idx >= len(row):
            return ""
        return cell_text(row[idx].value)

    known_cols = set(cols.values()) | {idx for idx, _ in hours_cols}
    warnings = []
    orphan_cols = set()

    categories = []
    current = None
    current_name = ""

    for row in ws.iter_rows(min_row=header_row + 1):
        cat = get(row, "category")
        task = get(row, "task")

        # 分類欄有值就開新區塊；沒值代表沿用上一個分類（合併儲存格或留白）。
        if cat and cat != current_name:
            current_name = cat
            current = {"category": cat, "items": [], "hours": {}}
            categories.append(current)

        if not task:
            continue

        if any(w in task for w in SUBTOTAL_WORDS):
            # 小計列：把工時掛回該分類，不當成作業項目。
            if current is not None:
                for idx, label in hours_cols:
                    val = cell_text(row[idx].value) if idx < len(row) else ""
                    if val:
                        current["hours"][label.strip()] = val
            continue

        if current is None:
            current = {"category": "(未分類)", "items": [], "hours": {}}
            categories.append(current)

        item = {
            "no": get(row, "no"),
            "task": task,
            "note": normalize_multiline(get(row, "note")),
            "acceptance": normalize_multiline(get(row, "acceptance")),
        }
        qty, unit = get(row, "qty"), get(row, "unit")
        if qty or unit:
            item["quantity"] = (qty + " " + unit).strip()
        item_hours = {}
        for idx, label in hours_cols:
            val = cell_text(row[idx].value) if idx < len(row) else ""
            if val and val != "0":
                item_hours[label.strip()] = val
        if item_hours:
            item["hours"] = item_hours

        # 標題列沒有涵蓋到的欄位：不要靜靜丟掉，收進 extra 並記下欄位代號。
        extra = []
        for idx, cell in enumerate(row):
            if idx in known_cols:
                continue
            text = cell_text(cell.value)
            if text:
                extra.append({"column": cell.column_letter, "value": normalize_multiline(text)})
                orphan_cols.add(cell.column_letter)
        if extra:
            item["extra"] = extra

        current["items"].append(item)

    if orphan_cols:
        warnings.append(
            "欄位 %s 有資料但標題列沒有對應欄名，內容已放進各項目的 extra；"
            "常見原因是打錯欄（例如驗收項目往右偏一欄），請人工判斷歸屬。"
            % "、".join(sorted(orphan_cols))
        )
    if "acceptance" not in cols:
        warnings.append("標題列沒有「驗收項目」欄，AGENTS.md 的驗收標準需另行向使用者確認。")

    categories = [c for c in categories if c["items"]]
    return {
        "source": path,
        "sheet": ws.title,
        "header_row": header_row,
        "columns": sorted(cols.keys()) + [l for _, l in hours_cols],
        "warnings": warnings,
        "categories": categories,
    }


def to_markdown(data):
    lines = [f"# 工時估算表解析：{data['source']}（工作表：{data['sheet']}）", ""]
    for w in data.get("warnings", []):
        lines.append(f"> ⚠️ {w}")
    if data.get("warnings"):
        lines.append("")
    for cat in data["categories"]:
        hours = "、".join(f"{k}={v}" for k, v in cat["hours"].items())
        lines.append(f"## {cat['category']}" + (f"　（小計 {hours}）" if hours else ""))
        for item in cat["items"]:
            no = f"{item['no']} " if item["no"] else ""
            lines.append(f"- {no}{item['task']}")
            for n in item["note"]:
                lines.append(f"  - 備註：{n}")
            for a in item["acceptance"]:
                lines.append(f"  - 驗收：{a}")
            for ex in item.get("extra", []):
                for v in ex["value"]:
                    lines.append(f"  - 未對應欄位({ex['column']})：{v}")
        lines.append("")
    return "\n".join(lines)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("path")
    ap.add_argument("--format", choices=["json", "md"], default="json")
    ap.add_argument("--sheet")
    ap.add_argument("--out", help="輸出檔路徑（UTF-8）。要接續處理時用這個，避免管線編碼問題。")
    args = ap.parse_args()

    data = parse(args.path, args.sheet)
    if args.format == "md":
        out = to_markdown(data)
    else:
        out = json.dumps(data, ensure_ascii=False, indent=2)

    if args.out:
        with open(args.out, "w", encoding="utf-8") as f:
            f.write(out + "\n")
        emit("wrote %s (%d categories)" % (args.out, len(data["categories"])), sys.stdout)
        for w in data.get("warnings", []):
            emit("WARNING: " + w)
        return

    # Windows 主控台預設是 cp950，中文會炸；直接寫 UTF-8 bytes 繞開。
    sys.stdout.buffer.write(out.encode("utf-8"))
    sys.stdout.buffer.write(b"\n")


if __name__ == "__main__":
    main()
